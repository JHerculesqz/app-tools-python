# -*- coding: utf-8 -*-

import shutil
from urllib import quote
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font


class ClientspiderPipeline(object):
    def __init__(self):
        self.workbook = XlsUtils.open('template.xlsx')
        self.sheet = XlsUtils.getSheet('Sheet', self.workbook)
        self.border = XlsUtils.getStyleBorder()
        self.align = XlsUtils.getStyleAlign()
        self.font = XlsUtils.getStyleFont()

    def process_item(self, item, spider):
        iRowIndex = XlsUtils.getRowCount(self.sheet) + 1
        # url
        oCell1 = XlsUtils.getCell(iRowIndex, 1, self.sheet)
        XlsUtils.setCell4Url('Link', item['url'], self.border, self.align, self.font, oCell1)
        # title
        oCell2 = XlsUtils.getCell(iRowIndex, 2, self.sheet)
        XlsUtils.setCell4Text(item['title'], self.border, self.align, self.font, oCell2)
        # title
        oCell3 = XlsUtils.getCell(iRowIndex, 3, self.sheet)
        XlsUtils.setCell4Text(item['time'], self.border, self.align, self.font, oCell3)
        # company
        oCell4 = XlsUtils.getCell(iRowIndex, 4, self.sheet)
        XlsUtils.setCell4Url(item['company'], 'https://www.qichacha.com/search?key=' + UrlUtils.encode(item['company']),
                             self.border, self.align, self.font, oCell4)
        # price
        oCell5 = XlsUtils.getCell(iRowIndex, 5, self.sheet)
        XlsUtils.setCell4Text(item['price'], self.border, self.align, self.font, oCell5)
        # phone
        oCell6 = XlsUtils.getCell(iRowIndex, 6, self.sheet)
        XlsUtils.setCell4Text(item['desc'], self.border, self.align, self.font, oCell6)
        # desc
        oCell7 = XlsUtils.getCell(iRowIndex, 7, self.sheet)
        XlsUtils.setCell4Text(item['phone'], self.border, self.align, self.font, oCell7)
        # save
        XlsUtils.save('result' + spider.name + '.xlsx', self.workbook)
        return item


# TODO:移到core-fw-python
class FileUtilsEx:
    def __init__(self):
        pass

    @staticmethod
    def copyFile(strFileSrc, strFileDst):
        shutil.copyfile(strFileSrc, strFileDst)


# TODO:移到core-fw-python
class XlsUtils:
    def __init__(self):
        pass

    @staticmethod
    def open(strFilePath):
        return load_workbook(strFilePath)

    @staticmethod
    def getSheet(strSheetName, oWorkbook):
        return oWorkbook[strSheetName]

    @staticmethod
    def getRowCount(oSheet):
        return oSheet.max_row

    @staticmethod
    def getStyleBorder():
        return Border(left=Side(border_style="thin", color='FF001000'),
                      right=Side(border_style="thin", color='FF110000'),
                      top=Side(border_style="thin", color='FF110000'),
                      bottom=Side(border_style="thin", color='FF110000'),
                      diagonal=Side(border_style=None, color='FF000000'),
                      diagonal_direction=0,
                      outline=Side(border_style=None, color='FF000000'),
                      vertical=Side(border_style=None, color='FF000000'),
                      horizontal=Side(border_style=None, color='FF110000'))

    @staticmethod
    def getStyleAlign():
        return Alignment(horizontal='center',
                         vertical='center',
                         text_rotation=0,
                         wrap_text=True,
                         shrink_to_fit=False,
                         indent=0)

    @staticmethod
    def getStyleFont():
        return Font(name=u'微软雅黑',
                    size=10,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    @staticmethod
    def getCell(iRowIndex, iColIndex, oSheet):
        return oSheet.cell(iRowIndex, iColIndex)

    @staticmethod
    def setCell4Url(strText, strUrl, oBorder, oAlign, oFont, oSheet):
        oSheet.value = strText
        oSheet.hyperlink = strUrl
        oSheet.border = oBorder
        oSheet.alignment = oAlign
        oSheet.font = oFont

    @staticmethod
    def setCell4Text(strText, oBorder, oAlign, oFont, oSheet):
        oSheet.value = strText
        oSheet.border = oBorder
        oSheet.alignment = oAlign
        oSheet.font = oFont

    @staticmethod
    def save(strFilePath, oWorkBook):
        oWorkBook.save(strFilePath)


# TODO:移到core-fw-python
class UrlUtils:
    def __init__(self):
        pass

    @staticmethod
    def encode(strUrl):
        return quote(strUrl)
